from django.shortcuts import render, redirect
from .models import InstitutionProfile
from student_app.models import StudentProfile
from Hr_app.models import HRProfile, Post_Job
from student_app.models import JobApplication
import csv
from django.contrib import messages
from django.core.mail import send_mail 
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import InstitutionProfile


def register(request):
    if request.method == "POST":
        institution_name = request.POST.get("institution_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        address = request.POST.get("address")

        # Check if email already exists
        if InstitutionProfile.objects.filter(email=email).exists():
            messages.error(request, "An institution with this email already exists.")
            return redirect("institution_register")

        # Save institution
        InstitutionProfile.objects.create(
            institution_name=institution_name,
            email=email,
            password=password,      # Use hashing in production
            address=address,
            status="Pending"
        )

        messages.success(
            request,
            "Registration successful! Your institution registration is awaiting admin approval."
        )

        return redirect("institution_login")

    return render(request, "institution_app/register.html")

def institution_login(request):

    if request.method == "POST":

        email = request.POST.get("email")
        password = request.POST.get("password")

        print("Email:", email)
        print("Password:", password)

        institution = InstitutionProfile.objects.filter(
            email=email,
            password=password
        ).first()

        print("Institution:", institution)

        if institution:

            request.session["institution_id"] = institution.id

            return redirect("institution_dashboard")

        return render(
            request,
            "institution_app/login.html",
            {
                "error": "Invalid Email or Password"
            }
        )

    return render(
        request,
        "institution_app/login.html"
    )


    



def institution_dashboard(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    institution = InstitutionProfile.objects.get(
        id=request.session["institution_id"]
    )

    total_students = StudentProfile.objects.count()

    total_companies = HRProfile.objects.count()

    total_jobs = Post_Job.objects.count()

    total_applications = JobApplication.objects.count()

    shortlisted = JobApplication.objects.filter(
        status="Shortlisted"
    ).count()

    placement_rate = 0

    if total_applications > 0:

        placement_rate = round(
            (shortlisted / total_applications) * 100
        )

    context = {

        "institution": institution,

        "total_students": total_students,

        "total_companies": total_companies,

        "total_jobs": total_jobs,

        "total_applications": total_applications,

        "shortlisted": shortlisted,

        "placement_rate": placement_rate

    }

    return render(
        request,
        "institution_app/dashboard.html",
        context
    )

def institution_students(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    students = StudentProfile.objects.all().order_by("name")

    return render(
        request,
        "institution_app/students.html",
        {
            "students": students
        }
    )
from Hr_app.models import HRProfile


def institution_companies(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    companies = HRProfile.objects.all().order_by("company_name")

    return render(
        request,
        "institution_app/companies.html",
        {
            "companies": companies
        }
    )
from django.shortcuts import render, redirect
from Hr_app.models import Post_Job


def institution_jobs(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    jobs = Post_Job.objects.all().order_by("-id")

    return render(
        request,
        "institution_app/jobs.html",
        {
            "jobs": jobs
        }
    )
from student_app.models import StudentProfile
from Hr_app.models import HRProfile, Post_Job
from student_app.models import JobApplication


def institution_analytics(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    total_students = StudentProfile.objects.count()

    total_companies = HRProfile.objects.count()

    total_jobs = Post_Job.objects.count()

    total_applications = JobApplication.objects.count()

    shortlisted = JobApplication.objects.filter(
        status="Shortlisted"
    ).count()

    selected = JobApplication.objects.filter(
        status="Selected"
    ).count()

    placement_rate = 0

    if total_students > 0:

        placement_rate = round(
            (selected / total_students) * 100
        )

    context = {

        "total_students": total_students,

        "total_companies": total_companies,

        "total_jobs": total_jobs,

        "total_applications": total_applications,

        "shortlisted": shortlisted,

        "selected": selected,

        "placement_rate": placement_rate

    }

    return render(
        request,
        "institution_app/analytics.html",
        context
    )
from student_app.models import StudentProfile
from Hr_app.models import HRProfile, Post_Job
from student_app.models import JobApplication


def institution_reports(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    context = {

        "students": StudentProfile.objects.count(),

        "companies": HRProfile.objects.count(),

        "jobs": Post_Job.objects.count(),

        "applications": JobApplication.objects.count(),

    }

    return render(
        request,
        "institution_app/reports.html",
        context
    )
from django.shortcuts import render, redirect, get_object_or_404
from .models import InstitutionProfile


def institution_profile(request):

    if "institution_id" not in request.session:
        return redirect("institution_login")

    institution = get_object_or_404(
        InstitutionProfile,
        id=request.session["institution_id"]
    )

    if request.method == "POST":

        institution.institution_name = request.POST.get(
            "institution_name"
        )

        institution.email = request.POST.get(
            "email"
        )

        institution.address = request.POST.get(
            "address"
        )

        institution.save()

        return redirect("institution_profile")

    return render(
        request,
        "institution_app/profile.html",
        {
            "institution": institution
        }
    )
def institution_logout(request):

    request.session.flush()

    return redirect("institution_login")
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from student_app.models import StudentProfile


def download_students_report(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Students Report"

    headers = [
        "Student Name",
        "Email",
        "Department",
        "CGPA",
        "Skills"
    ]

    fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col)

        cell.value = header

        cell.fill = fill

        cell.font = font

        cell.border = border

        cell.alignment = Alignment(horizontal="center")

    students = StudentProfile.objects.all()

    row = 2

    for student in students:

        ws.cell(row,1).value = student.name
        ws.cell(row,2).value = student.email
        ws.cell(row,3).value = student.department
        ws.cell(row,4).value = student.cgpa
        ws.cell(row,5).value = student.skills

        for col in range(1,6):

            ws.cell(row,col).border = border

        row += 1

    widths = {
        "A":25,
        "B":30,
        "C":20,
        "D":10,
        "E":35
    }

    for col,width in widths.items():

        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="Students_Report.xlsx"'

    wb.save(response)

    return response

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from Hr_app.models import HRProfile


def download_companies_report(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies Report"

    headers = [
        "Company Name",
        "HR Name",
        "Designation",
        "Email"
    ]

    fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Header Row
    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col)

        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    companies = HRProfile.objects.all()

    row = 2

    for company in companies:

        ws.cell(row,1).value = company.company_name
        ws.cell(row,2).value = company.name
        ws.cell(row,3).value = company.job
        ws.cell(row,4).value = company.email

        for col in range(1,5):

            ws.cell(row,col).border = border

        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Companies_Report.xlsx"'
    )

    wb.save(response)

    return response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from Hr_app.models import Post_Job


def download_jobs_report(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs Report"

    headers = [
        "Job Title",
        "Company",
        "Location",
        "Salary",
        "Required Skills",
        "Applicants"
    ]

    fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Header
    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col)

        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    jobs = Post_Job.objects.all()

    row = 2

    for job in jobs:

        ws.cell(row,1).value = job.job_title
        ws.cell(row,2).value = job.company.company_name
        ws.cell(row,3).value = job.location
        ws.cell(row,4).value = job.salary
        ws.cell(row,5).value = job.skills
        ws.cell(row,6).value = job.jobapplication_set.count()

        for col in range(1,7):
            ws.cell(row,col).border = border

        row += 1

    widths = {
        "A":30,
        "B":30,
        "C":20,
        "D":15,
        "E":40,
        "F":15
    }

    for col,width in widths.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Jobs_Report.xlsx"'
    )

    wb.save(response)

    return response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment



def download_applications_report(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications Report"

    headers = [
        "Student",
        "Company",
        "Job Title",
        "Status",
        "AI Score",
        "Recommendation"
    ]

    fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Header
    for col, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col)

        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    applications = JobApplication.objects.all()

    row = 2

    for app in applications:

        ws.cell(row,1).value = app.student.name
        ws.cell(row,2).value = app.job.company.company_name
        ws.cell(row,3).value = app.job.job_title
        ws.cell(row,4).value = app.status
        ws.cell(row,5).value = app.ai_score
        ws.cell(row,6).value = app.ai_recommendation

        for col in range(1,7):
            ws.cell(row,col).border = border

        row += 1

    widths = {
        "A":25,
        "B":30,
        "C":30,
        "D":18,
        "E":12,
        "F":20
    }

    for col,width in widths.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Placement_Report.xlsx"'
    )

    wb.save(response)

    return response

def institution_students(request):
    if 'institution_id' not in request.session: 
        return redirect('institution_login') 
    institution = InstitutionProfile.objects.get( id=request.session['institution_id'] ) 
    students = StudentProfile.objects.filter( college__iexact=institution.institution_name ).order_by('-id') 
    return render( request, 'Institution_app/students.html', { 'institution': institution, 'students': students } )

def approve_student(request, id): 
    if 'institution_id' not in request.session:
        return redirect('institution_login') 
    student = StudentProfile.objects.get(id=id) 
    student.status = 'Approved' 
    student.save() # Login URL 
    login_url = request.build_absolute_uri('/student/login/') 
    # Send approval email 
    send_mail( 
        subject='RecruitIQ Account Approved', 
        message=f''' 
    Hello {student.name}, 
    Congratulations! Your account has been approved by your institution. 
    You can now log in to RecruitIQ using the link below: 
    {login_url} 
    Best regards,
    RecruitIQ Team 
    ''', 
        from_email=settings.EMAIL_HOST_USER, 
        recipient_list=[student.email], 
        fail_silently=False, ) 
    return redirect('institution_students')

def reject_student(request, id):
    if 'institution_id' not in request.session: 
        return redirect('institution_login') 
    student = StudentProfile.objects.get(id=id) 
    student.status = 'Rejected' 
    student.save() 
    return redirect('institution_students')